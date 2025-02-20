import time
from validate_email import validate_email
from openpyxl import Workbook, load_workbook



log_or_sing = str(input("Do you want to login or singup: "))


####### Setting up Excel sheet: Workbook & Worksheet 👇 ######
workbook = load_workbook('Login_System.xlsx')
info_sheet = workbook['Username_Password']



######### Start: Timer 👇 #####
def countdown_30():
    countdown_second = 30
    for i in range(countdown_second, 0, -1):
        print(f"Time remaining: {i} seconds", end='\r')
        time.sleep(1)


def countdown_ten():
    countdown_second = 10
    for i in range(countdown_second, 0, -1):
        print(f"Time remaining: {i} seconds", end='\r')
        time.sleep(1)
######### End: Timer 👆 ########


####### Start: Singup 👇 ########
def singup():
    email = input('Enter your Email: ')

    ##### Start: Email Checker 👇 ######
    mail_check = validate_email(email, verify=True)

    while mail_check == False:
        print('Invalid Email. Try Again.')
        email = input('Enter Email: ')
        mail_check = validate_email(email, verify=True)
    ##### End: Email Checker 👆 ######

    username = input('Enter username for singup: ')
    password = input('Enter password for singup: ')
    confirm_password = input('Confirm Password for singup: ')

    try_time = 0

    if confirm_password == password:
        print('Singup Successfull.')
        #### Updating credentials to the excel sheet 👇 #####
        info_sheet['A1'] = "Email"
        info_sheet['B1'] = "Username"
        info_sheet['C1'] = "Password"

        # Find the first truly empty row by scanning through the first column
        for row in range(1, info_sheet.max_row + 1):
            if info_sheet.cell(row=row, column=1).value is None:
                next_row = row
                break
        else:
            # If no empty cell is found, append the row after the last filled row
            next_row = info_sheet.max_row + 1

        info_sheet.cell(row=next_row, column=1, value=email)
        info_sheet.cell(row=next_row, column=2, value=username)
        info_sheet.cell(row=next_row, column=3, value=password)
        workbook.save('Login_System.xlsx')


    elif password != confirm_password and try_time < 5:
        while password != confirm_password and try_time < 5:
            print("Password didn't match. Try again!")
            confirm_password = input("Confirm you password: ")
            try_time = try_time + 1
        
            if try_time == 5:
                print("Too many attempts. Try after 10 second.")
                time_to_wait = countdown_ten() #time.sleep(10)

                while password != confirm_password and try_time < 8:
                    confirm_password = input('Confirm Password: ')
                    try_time = try_time + 1

                    if try_time == 8:
                        print('Again Too many attempts. Try after 30 Seconds.')
                        time_to_wait = countdown_30() #time.sleep(30)

                        while password != confirm_password and try_time < 10:
                            confirm_password = input('Confirm Password: ')
                            try_time = try_time + 1

                        if try_time == 10:
                            print("Attempt finised. Can't try anymore.")
####### End: Singup 👆 ########


####### Start: Login 👇 ########
def login():

    email = input('Enter Email: ')

    ##### Start: Email Checker 👇 ######
    mail_check = validate_email(email, verify=True)

    while mail_check == False:
        print('Invalid Email. Try Again.')
        email = input('Enter Email: ')
        mail_check = validate_email(email, verify=True)
    ##### End: Email Checker 👆 ######

    email_value = False
    for email_cell in info_sheet['A']:
        if email_cell.value == email:
            email_value = True
            break


    username = input('Enter Username: ')   

    username_value = False
    for username_cell in info_sheet['B']:
        if username_cell.value == username:
            username_value = True
            break


    password = input('Enter Password: ')
    
    password_value = False
    for password_cell in info_sheet['C']:
        if password_cell.value == password:
            password_value = True
            break

    
    if email_value and username_value and password_value:
        print('Login Successful!')
    
    else:
        print("Username can't recognise. Try Singup.")
        print(singup())
        
####### End: Login 👆 ########

workbook.save('Login_System.xlsx')

if log_or_sing == 'login':
    print(login())
elif log_or_sing == 'singup':
    print(singup())
else:
    print('Enter only login or singup.')
